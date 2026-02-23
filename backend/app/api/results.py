import csv
import io
import logging
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import distinct, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from app.models.job import Job
from app.models.lead import Lead
from app.schemas import (
    JobStatsResponse,
    LeadBatchActionResponse,
    LeadBatchDeleteRequest,
    LeadBatchUpdateRequest,
    LeadListResponse,
    LeadResponse,
    LeadUpdate,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/results", tags=["results"])

SORTABLE_COLUMNS = {
    "name": Lead.name,
    "rating": Lead.rating,
    "review_count": Lead.review_count,
    "created_at": Lead.created_at,
}


# ---------------------------------------------------------------------------
# Lead list (enhanced with favorites, archive, tags, search)
# ---------------------------------------------------------------------------

@router.get(
    "/{job_id}",
    response_model=LeadListResponse,
    summary="Get job results",
    description="Paginated, sortable, filterable leads for a scrape job.",
)
async def get_results(
    job_id: uuid.UUID,
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    sort_by: str = Query(
        default="name",
        pattern="^(name|rating|review_count|created_at)$",
    ),
    sort_order: str = Query(default="asc", pattern="^(asc|desc)$"),
    search: str | None = Query(default=None, max_length=200),
    has_phone: bool | None = Query(default=None),
    has_website: bool | None = Query(default=None),
    min_rating: float | None = Query(default=None, ge=0, le=5),
    favorite_only: bool = Query(default=False),
    archived: bool = Query(default=False),
    tags: str | None = Query(default=None, max_length=500),
    db: AsyncSession = Depends(get_db),
) -> dict:
    job_result = await db.execute(
        select(Job.id).where(Job.id == job_id),
    )
    if not job_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Job not found")

    query = select(Lead).where(Lead.job_id == job_id)

    # Filter: favorites
    if favorite_only:
        query = query.where(Lead.is_favorite.is_(True))

    # Filter: archived (default=False → exclude archived)
    if not archived:
        query = query.where(Lead.is_archived.is_(False))

    # Filter: search across name, address, phone, email
    if search:
        pattern = f"%{search}%"
        query = query.where(
            or_(
                Lead.name.ilike(pattern),
                Lead.address.ilike(pattern),
                Lead.phone.ilike(pattern),
                Lead.primary_email.ilike(pattern),
            )
        )

    # Filter: tags (comma-separated)
    if tags:
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]
        for tag in tag_list:
            query = query.where(
                Lead.tags.cast(str).ilike(f"%{tag}%"),
            )

    if has_phone is True:
        query = query.where(Lead.phone.is_not(None))
    elif has_phone is False:
        query = query.where(Lead.phone.is_(None))
    if has_website is True:
        query = query.where(Lead.website.is_not(None))
    elif has_website is False:
        query = query.where(Lead.website.is_(None))
    if min_rating is not None:
        query = query.where(Lead.rating >= min_rating)

    # Count
    count_query = select(func.count()).select_from(query.subquery())
    count_result = await db.execute(count_query)
    total = count_result.scalar_one()

    # Sort
    sort_col = SORTABLE_COLUMNS.get(sort_by, Lead.name)
    order = sort_col.desc() if sort_order == "desc" else sort_col.asc()
    query = query.order_by(order.nulls_last()).offset(skip).limit(limit)

    result = await db.execute(query)
    items = list(result.scalars().all())

    return {"items": items, "total": total, "skip": skip, "limit": limit}


# ---------------------------------------------------------------------------
# Single lead CRUD
# ---------------------------------------------------------------------------

@router.patch(
    "/{lead_id}",
    response_model=LeadResponse,
    summary="Update a lead",
    description="Update user-managed fields (favorite, archived, notes, tags).",
)
async def update_lead(
    lead_id: uuid.UUID,
    payload: LeadUpdate,
    db: AsyncSession = Depends(get_db),
) -> Lead:
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id),
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(lead, key, value)

    await db.flush()
    await db.refresh(lead)
    return lead


@router.delete(
    "/{lead_id}",
    status_code=204,
    summary="Delete a lead",
)
async def delete_lead(
    lead_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
) -> None:
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id),
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    await db.delete(lead)
    await db.flush()


# ---------------------------------------------------------------------------
# Bulk lead operations
# ---------------------------------------------------------------------------

@router.post(
    "/batch-update",
    response_model=LeadBatchActionResponse,
    summary="Bulk update leads",
)
async def batch_update_leads(
    payload: LeadBatchUpdateRequest,
    db: AsyncSession = Depends(get_db),
) -> dict:
    update_data = payload.update.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(
            status_code=400, detail="No fields to update",
        )

    result = await db.execute(
        update(Lead)
        .where(Lead.id.in_(payload.lead_ids))
        .values(**update_data)
    )
    await db.flush()
    return {"count": result.rowcount}


@router.post(
    "/batch-delete",
    response_model=LeadBatchActionResponse,
    summary="Bulk delete leads",
)
async def batch_delete_leads(
    payload: LeadBatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
) -> dict:
    result = await db.execute(
        select(Lead).where(Lead.id.in_(payload.lead_ids)),
    )
    leads = result.scalars().all()
    count = 0
    for lead in leads:
        await db.delete(lead)
        count += 1
    await db.flush()
    return {"count": count}


# ---------------------------------------------------------------------------
# Export: CSV + Excel
# ---------------------------------------------------------------------------

CSV_TEMPLATES = {
    "default": {
        "fieldnames": [
            "place_id", "name", "address", "phone", "website",
            "primary_email", "rating", "review_count", "business_type",
            "types", "latitude", "longitude", "price_level",
            "business_status", "maps_url", "description", "verified",
            "owner_name", "employee_count", "year_established",
            "business_age_years", "social_facebook", "social_linkedin",
            "social_instagram", "social_twitter", "social_youtube",
            "source",
        ],
    },
    "clay": {
        "fieldnames": [
            "Company Name", "Website", "Phone Number", "Email",
            "Address", "Rating", "Reviews", "Contact Name",
            "Company Size", "Founded Year", "LinkedIn URL",
            "Facebook URL", "Instagram URL", "Company Description",
        ],
    },
    "hubspot": {
        "fieldnames": [
            "Company name", "Company domain", "Phone number",
            "Address", "City", "State", "Zip", "Country",
            "Number of employees", "Year founded", "Description",
        ],
    },
    "outreach": {
        "fieldnames": [
            "Company", "Contact Name", "Email", "Phone",
            "Website", "LinkedIn", "Rating", "Notes",
        ],
    },
}


def _lead_to_default_row(lead: Lead) -> dict:
    social = lead.social_links or {}
    return {
        "place_id": lead.place_id,
        "name": lead.name,
        "address": lead.address or "",
        "phone": lead.phone or "",
        "website": lead.website or "",
        "primary_email": lead.primary_email or "",
        "rating": lead.rating or "",
        "review_count": lead.review_count or "",
        "business_type": lead.business_type or "",
        "types": ",".join(lead.types) if lead.types else "",
        "latitude": lead.latitude or "",
        "longitude": lead.longitude or "",
        "price_level": lead.price_level or "",
        "business_status": lead.business_status or "",
        "maps_url": lead.maps_url or "",
        "description": lead.description or "",
        "verified": (
            lead.verified if lead.verified is not None else ""
        ),
        "owner_name": lead.owner_name or "",
        "employee_count": lead.employee_count or "",
        "year_established": lead.year_established or "",
        "business_age_years": lead.business_age_years or "",
        "social_facebook": social.get("facebook", ""),
        "social_linkedin": social.get("linkedin", ""),
        "social_instagram": social.get("instagram", ""),
        "social_twitter": social.get("twitter", ""),
        "social_youtube": social.get("youtube", ""),
        "source": lead.source,
    }


def _lead_to_clay_row(lead: Lead) -> dict:
    social = lead.social_links or {}
    return {
        "Company Name": lead.name,
        "Website": lead.website or "",
        "Phone Number": lead.phone or "",
        "Email": lead.primary_email or "",
        "Address": lead.address or "",
        "Rating": lead.rating or "",
        "Reviews": lead.review_count or "",
        "Contact Name": lead.owner_name or "",
        "Company Size": lead.employee_count or "",
        "Founded Year": lead.year_established or "",
        "LinkedIn URL": social.get("linkedin", ""),
        "Facebook URL": social.get("facebook", ""),
        "Instagram URL": social.get("instagram", ""),
        "Company Description": lead.description or "",
    }


def _lead_to_hubspot_row(lead: Lead) -> dict:
    address = lead.address or ""
    parts = [p.strip() for p in address.split(",")]
    city = parts[-3] if len(parts) >= 3 else ""
    state = parts[-2] if len(parts) >= 2 else ""
    country = parts[-1] if len(parts) >= 1 else ""
    domain = ""
    if lead.website:
        try:
            from urllib.parse import urlparse
            domain = urlparse(lead.website).netloc
        except Exception:
            domain = lead.website
    return {
        "Company name": lead.name,
        "Company domain": domain,
        "Phone number": lead.phone or "",
        "Address": address,
        "City": city,
        "State": state,
        "Zip": "",
        "Country": country,
        "Number of employees": lead.employee_count or "",
        "Year founded": lead.year_established or "",
        "Description": lead.description or "",
    }


def _lead_to_outreach_row(lead: Lead) -> dict:
    social = lead.social_links or {}
    return {
        "Company": lead.name,
        "Contact Name": lead.owner_name or "",
        "Email": lead.primary_email or "",
        "Phone": lead.phone or "",
        "Website": lead.website or "",
        "LinkedIn": social.get("linkedin", ""),
        "Rating": lead.rating or "",
        "Notes": lead.notes or "",
    }


TEMPLATE_ROW_FN = {
    "default": _lead_to_default_row,
    "clay": _lead_to_clay_row,
    "hubspot": _lead_to_hubspot_row,
    "outreach": _lead_to_outreach_row,
}


@router.get(
    "/{job_id}/export",
    summary="Export results as CSV or Excel",
    description=(
        "Download leads. format=csv|xlsx, "
        "template=default|clay|hubspot|outreach."
    ),
)
async def export_results(
    job_id: uuid.UUID,
    template: str = Query(
        default="default",
        pattern="^(default|clay|hubspot|outreach)$",
    ),
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    favorites_only: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    job_result = await db.execute(
        select(Job).where(Job.id == job_id),
    )
    job = job_result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    query = select(Lead).where(Lead.job_id == job_id)
    if favorites_only:
        query = query.where(Lead.is_favorite.is_(True))
    query = query.order_by(Lead.name)

    result = await db.execute(query)
    leads = result.scalars().all()

    row_fn = TEMPLATE_ROW_FN[template]
    tmpl = CSV_TEMPLATES[template]

    if format == "xlsx":
        return _export_xlsx(
            leads, row_fn, tmpl["fieldnames"], job, template,
        )

    # CSV export
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=tmpl["fieldnames"])
    writer.writeheader()
    for lead in leads:
        writer.writerow(row_fn(lead))

    output.seek(0)
    filename = f"leads_{job_id}_{template}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _export_xlsx(
    leads: list,
    row_fn,
    fieldnames: list[str],
    job,
    template: str,
) -> StreamingResponse:
    """Generate a formatted Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed",
        )

    wb = Workbook()
    ws = wb.active
    sheet_name = f"Leads - {job.keyword} in {job.location}"
    ws.title = sheet_name[:31]

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    highlight_fill = PatternFill("solid", fgColor="FEF9C3")

    for row_idx, lead in enumerate(leads, 2):
        row_data = row_fn(lead)
        for col_idx, field in enumerate(fieldnames, 1):
            value = row_data.get(field, "")
            cell = ws.cell(
                row=row_idx, column=col_idx, value=value,
            )
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if field in ("rating", "Rating"):
                try:
                    if float(value) >= 4.5:
                        cell.fill = highlight_fill
                except (ValueError, TypeError):
                    pass

    # Auto-size columns
    for col_idx, field in enumerate(fieldnames, 1):
        max_len = len(str(field))
        for row in ws.iter_rows(
            min_row=2, max_row=min(len(leads) + 1, 50),
            min_col=col_idx, max_col=col_idx,
        ):
            for cell in row:
                val_len = len(str(cell.value or ""))
                if val_len > max_len:
                    max_len = val_len
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"leads_{job.id}_{template}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ---------------------------------------------------------------------------
# Job-level stats
# ---------------------------------------------------------------------------

@router.get(
    "/{job_id}/stats",
    response_model=JobStatsResponse,
    summary="Get job statistics",
)
async def get_stats(
    job_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
) -> dict:
    job_result = await db.execute(
        select(Job.id).where(Job.id == job_id),
    )
    if not job_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Job not found")

    total_result = await db.execute(
        select(func.count())
        .where(Lead.job_id == job_id),
    )
    total_leads = total_result.scalar_one()

    unique_result = await db.execute(
        select(func.count(distinct(Lead.place_id)))
        .where(Lead.job_id == job_id),
    )
    unique_place_ids = unique_result.scalar_one()

    source_result = await db.execute(
        select(Lead.source, func.count())
        .where(Lead.job_id == job_id)
        .group_by(Lead.source),
    )
    sources = dict(source_result.all())

    avg_result = await db.execute(
        select(func.avg(Lead.rating))
        .where(Lead.job_id == job_id)
        .where(Lead.rating.is_not(None)),
    )
    avg_raw = avg_result.scalar_one()
    avg_rating = round(float(avg_raw), 2) if avg_raw else None

    phone_result = await db.execute(
        select(func.count())
        .where(Lead.job_id == job_id)
        .where(Lead.phone.is_not(None)),
    )
    with_phone = phone_result.scalar_one()

    website_result = await db.execute(
        select(func.count())
        .where(Lead.job_id == job_id)
        .where(Lead.website.is_not(None)),
    )
    with_website = website_result.scalar_one()

    email_result = await db.execute(
        select(func.count())
        .where(Lead.job_id == job_id)
        .where(Lead.primary_email.is_not(None)),
    )
    with_email = email_result.scalar_one()

    type_result = await db.execute(
        select(Lead.business_type, func.count())
        .where(Lead.job_id == job_id)
        .where(Lead.business_type.is_not(None))
        .group_by(Lead.business_type)
        .order_by(func.count().desc())
        .limit(20),
    )
    business_types = dict(type_result.all())

    return {
        "job_id": job_id,
        "total_leads": total_leads,
        "unique_place_ids": unique_place_ids,
        "sources": sources,
        "avg_rating": avg_rating,
        "with_phone": with_phone,
        "with_website": with_website,
        "with_email": with_email,
        "business_types": business_types,
    }
